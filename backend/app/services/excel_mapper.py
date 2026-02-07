"""
Excel Mapper Module
Converts Excel/CSV files to Azure Document Intelligence compatible JSON format.
This allows Excel files to be processed through the same LLM extraction pipeline
without calling Azure OCR services.

Beta Feature: use_virtual_excel_ocr
"""
import io
import logging
from typing import Dict, Any, List, Optional, Union
import aiohttp

logger = logging.getLogger(__name__)

# Virtual canvas dimensions
VIRTUAL_WIDTH = 1000
CELL_HEIGHT = 50  # Height per row


class ExcelMapper:
    """
    Maps Excel/CSV data to Azure Document Intelligence JSON format.
    """
    
    @classmethod
    async def from_url(cls, file_url: str) -> Dict[str, Any]:
        """
        Download file from URL and convert to Doc Intel format.
        """
        try:
            # Download file
            async with aiohttp.ClientSession() as session:
                async with session.get(file_url) as response:
                    if response.status != 200:
                        raise ValueError(f"Failed to download file: {response.status}")
                    file_bytes = await response.read()
            
            # Determine file type from URL (strip SAS query params)
            ext = file_url.rsplit(".", 1)[-1].lower().split("?")[0]
            
            if ext == "csv":
                return cls._parse_csv(file_bytes)
            elif ext in ("xlsx", "xls"):
                return cls._parse_excel(file_bytes)
            else:
                raise ValueError(f"Unsupported file extension: {ext}")
                
        except Exception as e:
            logger.error(f"[ExcelMapper] Error processing file: {e}")
            raise e
    
    @classmethod
    def from_bytes(cls, file_bytes: bytes, file_type: str = "xlsx") -> Dict[str, Any]:
        """
        Convert file bytes to Doc Intel format.
        """
        if file_type == "csv":
            return cls._parse_csv(file_bytes)
        else:
            return cls._parse_excel(file_bytes)
    
    @classmethod
    def _parse_csv(cls, file_bytes: bytes) -> Dict[str, Any]:
        """
        Parse CSV file to Doc Intel format.
        """
        import csv
        
        # Try to decode with different encodings
        content_str = None
        for encoding in ["utf-8", "cp949", "euc-kr", "latin-1"]:
            try:
                content_str = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if content_str is None:
            raise ValueError("Failed to decode CSV file with supported encodings")
        
        # Parse CSV
        reader = csv.reader(io.StringIO(content_str))
        rows = list(reader)
        
        if not rows:
            return cls._empty_result()
        
        return cls._rows_to_doc_intel(rows, sheet_name="Sheet1", page_number=1)
    
    @classmethod
    def _parse_excel(cls, file_bytes: bytes) -> Dict[str, Any]:
        """
        Parse Excel file to Doc Intel format.
        Each sheet becomes a separate page.
        """
        try:
            import openpyxl
        except ImportError:
            logger.warning("[ExcelMapper] openpyxl not available, falling back to pandas")
            return cls._parse_excel_pandas(file_bytes)
        
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        all_pages = []
        all_tables = []
        all_content_parts = []
        
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            page_number = sheet_idx + 1
            
            # Extract rows
            rows = []
            for row in sheet.iter_rows():
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                rows.append(row_values)
            
            if not rows:
                continue
            
            # Build page and table for this sheet
            result = cls._rows_to_doc_intel(rows, sheet_name, page_number)
            
            all_pages.extend(result["pages"])
            all_tables.extend(result["tables"])
            all_content_parts.append(result["content"])
        
        return {
            "content": "\n\n".join(all_content_parts),
            "pages": all_pages,
            "tables": all_tables,
            "paragraphs": [],
            "key_value_pairs": [],
            "documents": [],
            "_layout_parser_bypass": True  # Explicitly tell pipeline to skip LayoutParser
        }
    
    @classmethod
    def _parse_excel_pandas(cls, file_bytes: bytes) -> Dict[str, Any]:
        """
        Fallback Excel parser using pandas.
        """
        import pandas as pd
        
        excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
        
        all_pages = []
        all_tables = []
        all_content_parts = []
        
        for sheet_idx, sheet_name in enumerate(excel_file.sheet_names):
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
            page_number = sheet_idx + 1
            
            # Convert DataFrame to list of rows
            rows = df.fillna("").astype(str).values.tolist()
            
            if not rows:
                continue
            
            result = cls._rows_to_doc_intel(rows, sheet_name, page_number)
            
            all_pages.extend(result["pages"])
            all_tables.extend(result["tables"])
            all_content_parts.append(result["content"])
        
        return {
            "content": "\n\n".join(all_content_parts),
            "pages": all_pages,
            "tables": all_tables,
            "paragraphs": [],
            "key_value_pairs": [],
            "documents": []
        }
    
    @classmethod
    def _rows_to_doc_intel(cls, rows: List[List[str]], sheet_name: str, page_number: int) -> Dict[str, Any]:
        """
        Convert a 2D array of rows to Doc Intel format.
        """
        if not rows:
            return cls._empty_result()
        
        row_count = len(rows)
        col_count = max(len(row) for row in rows) if rows else 0
        
        if col_count == 0:
            return cls._empty_result()
        
        # Virtual dimensions
        page_width = VIRTUAL_WIDTH
        page_height = row_count * CELL_HEIGHT
        cell_width = page_width / col_count
        
        # Build content string
        content_lines = []
        for row in rows:
            content_lines.append("\t".join(row))
        content = "\n".join(content_lines)
        
        # Build words (each cell as a word)
        words = []
        lines = []
        
        for row_idx, row in enumerate(rows):
            line_content_parts = []
            for col_idx, cell_value in enumerate(row):
                if not cell_value.strip():
                    continue
                
                # Calculate virtual polygon
                x1 = col_idx * cell_width
                y1 = row_idx * CELL_HEIGHT
                x2 = (col_idx + 1) * cell_width
                y2 = (row_idx + 1) * CELL_HEIGHT
                
                # 4-point polygon [x1,y1, x2,y1, x2,y2, x1,y2]
                polygon = [x1, y1, x2, y1, x2, y2, x1, y2]
                
                words.append({
                    "content": cell_value,
                    "polygon": polygon,
                    "confidence": 1.0  # Virtual OCR is "perfect"
                })
                
                line_content_parts.append(cell_value)
            
            if line_content_parts:
                # Line spans entire row
                line_y1 = row_idx * CELL_HEIGHT
                line_y2 = (row_idx + 1) * CELL_HEIGHT
                lines.append({
                    "content": "\t".join(line_content_parts),
                    "polygon": [0, line_y1, page_width, line_y1, page_width, line_y2, 0, line_y2]
                })
        
        # Build table cells
        cells = []
        for row_idx, row in enumerate(rows):
            for col_idx, cell_value in enumerate(row):
                x1 = col_idx * cell_width
                y1 = row_idx * CELL_HEIGHT
                x2 = (col_idx + 1) * cell_width
                y2 = (row_idx + 1) * CELL_HEIGHT
                
                polygon = [x1, y1, x2, y1, x2, y2, x1, y2]
                
                cells.append({
                    "row_index": row_idx,
                    "column_index": col_idx,
                    "content": cell_value,
                    "kind": "columnHeader" if row_idx == 0 else "content",
                    "bounding_regions": [{
                        "page_number": page_number,
                        "polygon": polygon
                    }]
                })
        
        # Build page object
        page = {
            "page_number": page_number,
            "width": page_width,
            "height": page_height,
            "unit": "pixel",  # Virtual pixels
            "words": words,
            "lines": lines,
            "selection_marks": []
        }
        
        # Build table object
        table = {
            "row_count": row_count,
            "column_count": col_count,
            "cells": cells,
            "bounding_regions": [{
                "page_number": page_number,
                "polygon": [0, 0, page_width, 0, page_width, page_height, 0, page_height]
            }]
        }
        
        return {
            "content": content,
            "pages": [page],
            "tables": [table],
            "paragraphs": [],
            "key_value_pairs": [],
            "documents": []
        }
    
    @classmethod
    def _empty_result(cls) -> Dict[str, Any]:
        """
        Return empty Doc Intel format.
        """
        return {
            "content": "",
            "pages": [],
            "tables": [],
            "paragraphs": [],
            "key_value_pairs": [],
            "documents": []
        }
